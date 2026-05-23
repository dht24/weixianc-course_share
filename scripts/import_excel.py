from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
LEGACY_EXCEL = Path(r"C:\Users\dht\Downloads\选课分享.xlsx")
TIC2_EXCEL = Path(r"C:\Users\dht\Downloads\TIC2选课.xlsx")
DEFAULT_EXCEL_FILES = [LEGACY_EXCEL, TIC2_EXCEL]
OUT = ROOT / "src" / "data" / "seed-data.ts"
SQL_OUT = ROOT / "supabase" / "seed.sql"
UPSERT_SQL_OUT = ROOT / "supabase" / "upsert-seed.sql"


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def slugify(text: str) -> str:
    normalized = re.sub(r"\s+", "-", text.strip().lower())
    normalized = re.sub(r"[^\w\u4e00-\u9fff（）()\-]+", "", normalized)
    return normalized.strip("-") or "item"


def detect_workload(text: str) -> str:
    if any(word in text for word in ["作业量很大", "课程量大", "任务量大", "偏多", "很大"]):
        return "重"
    if any(word in text for word in ["作业不多", "任务量小", "课外时间", "轻松", "不占用"]):
        return "轻"
    return "不确定"


def detect_grading(text: str) -> str:
    if any(word in text for word in ["给分好", "好拿高分", "包4.0", "A+", "给分不错", "给分也好"]):
        return "偏好"
    if any(word in text for word in ["给分差", "给分都很差", "难拿高分", "偏严"]):
        return "偏严"
    if "给分" in text:
        return "正常"
    return "不确定"


def detect_assessment(text: str) -> str:
    parts: list[str] = []
    for label, words in {
        "考试": ["考试", "期中", "期末", "闭卷", "开卷"],
        "论文": ["论文", "小论文", "大论文"],
        "展示": ["pre", "展示", "汇报"],
        "作业": ["作业", "报告"],
    }.items():
        if any(word in text for word in words):
            parts.append(label)
    return " / ".join(parts) if parts else "不确定"


def score(text: str) -> int:
    if any(word in text for word in ["神", "无脑入", "强力推荐", "推荐", "很好", "非常好"]):
        return 5
    if any(word in text for word in ["不推荐", "差评", "给分差"]):
        return 2
    if any(word in text for word in ["中规中矩", "还行", "一般"]):
        return 3
    return 4


COURSE_NAME_MAP = {
    "24770031 (2A)": "科技创新与挑战2A",
    "24770041 (2B)": "科技创新与挑战2B",
    "24770051 (2C)": "科技创新与挑战2C",
    "24770061 (2D)": "科技创新与挑战2D",
    "2C": "科技创新与挑战2C",
}

OBSOLETE_TIC2_COURSE_NAMES = [
    "课程号",
    "24770031 (2A)",
    "24770041 (2B)",
    "24770051 (2C)",
    "24770061 (2D)",
    "2C",
]


def normalize_course_name(value: str) -> str:
    return COURSE_NAME_MAP.get(value, value)


def read_tic2_excel(path: Path) -> dict[str, dict[str, list[str]]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    courses: dict[str, dict[str, list[str]]] = defaultdict(lambda: defaultdict(list))
    current_course = ""
    current_teacher = ""

    for row in range(2, ws.max_row + 1):
        raw_course = clean(ws.cell(row, 1).value)
        teacher = clean(ws.cell(row, 3).value)
        if raw_course:
            current_course = normalize_course_name(raw_course)
        if teacher:
            current_teacher = teacher
            courses[current_course][current_teacher]

        if not current_course or not current_teacher:
            continue

        review_values = [clean(ws.cell(row, col).value) for col in range(6, ws.max_column + 1)]

        # The source sheet has a final 2C note placed under the research-direction column.
        if current_course == "科技创新与挑战2C" and not any(review_values):
            extra_note = clean(ws.cell(row, 5).value)
            if row > 1 and not teacher and extra_note:
                review_values.append(extra_note)

        for review in review_values:
            if review:
                courses[current_course][current_teacher].append(review)

    courses.pop("课程号", None)
    return courses


def read_legacy_excel(path: Path) -> dict[str, dict[str, list[str]]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    merged: dict[tuple[int, int], Any] = {}
    for cell_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = cell_range.bounds
        merged_value = ws.cell(min_row, min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged[(row, col)] = merged_value

    def value(row: int, col: int) -> str:
        return clean(ws.cell(row, col).value or merged.get((row, col)))

    courses: dict[str, dict[str, list[str]]] = defaultdict(lambda: defaultdict(list))
    current_course = ""
    current_teacher = ""

    for row in range(4, ws.max_row + 1):
        course = value(row, 1)
        teacher = value(row, 2)
        if course:
            current_course = course
            current_teacher = ""
        if teacher:
            current_teacher = teacher
        if not current_course or not current_teacher:
            continue

        for col in range(3, ws.max_column + 1):
            review = value(row, col)
            if review:
                courses[current_course][current_teacher].append(review)

    return courses


def read_excel(path: Path) -> dict[str, dict[str, list[str]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header = [clean(cell) for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    wb.close()
    if len(header) >= 3 and header[0] == "课程号" and header[2] == "授课教师":
        return read_tic2_excel(path)
    return read_legacy_excel(path)


def merge_raw_data(paths: list[Path]) -> dict[str, dict[str, list[str]]]:
    merged: dict[str, dict[str, list[str]]] = defaultdict(lambda: defaultdict(list))
    for path in paths:
        raw = read_excel(path)
        for course_name, teacher_map in raw.items():
            if course_name in OBSOLETE_TIC2_COURSE_NAMES:
                continue
            for teacher_name, reviews in teacher_map.items():
                merged[course_name][teacher_name].extend(reviews)
    return merged


def build_seed(raw: dict[str, dict[str, list[str]]]) -> dict[str, Any]:
    courses = []
    teachers_by_name: dict[str, dict[str, str]] = {}
    offerings = []
    reviews = []

    for course_index, (course_name, teacher_map) in enumerate(raw.items(), start=1):
        course_id = f"course-{course_index:03d}-{slugify(course_name)}"
        courses.append(
            {
                "id": course_id,
                "name": course_name,
                "category": "",
                "audience": "",
                "program": "课程概要",
                "aliases": [],
                "summary": "",
            }
        )

        for teacher_name, teacher_reviews in teacher_map.items():
            if teacher_name not in teachers_by_name:
                teacher_id = f"teacher-{len(teachers_by_name) + 1:03d}-{slugify(teacher_name)[:24]}"
                teachers_by_name[teacher_name] = {
                    "id": teacher_id,
                    "name": teacher_name,
                    "department": "",
                    "note": "",
                }
            teacher = teachers_by_name[teacher_name]
            offering_id = f"offering-{len(offerings) + 1:03d}"
            offerings.append(
                {
                    "id": offering_id,
                    "courseId": course_id,
                    "teacherId": teacher["id"],
                    "term": "待补充",
                    "status": "published",
                }
            )

            for review_text in teacher_reviews:
                reviews.append(
                    {
                        "id": f"review-{len(reviews) + 1:04d}",
                        "offeringId": offering_id,
                        "term": "",
                        "workload": "不确定",
                        "grading": "不确定",
                        "assessment": "不确定",
                        "rating": 4,
                        "content": review_text,
                        "tags": [],
                        "contact": "",
                        "status": "published",
                        "createdAt": "2026-05-22",
                    }
                )

    return {
        "courses": courses,
        "teachers": list(teachers_by_name.values()),
        "offerings": offerings,
        "reviews": reviews,
        "resources": [],
    }


def sql_quote(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, list):
        return "array[" + ", ".join(sql_quote(item) for item in value) + "]::text[]"
    text = str(value).replace("'", "''")
    return f"'{text}'"


def insert_statement(table: str, columns: list[str], values: list[Any], conflict: str | None = None) -> str:
    base = f"insert into {table} ({', '.join(columns)}) values ({', '.join(sql_quote(value) for value in values)})"
    if conflict:
        update_columns = [column for column in columns if column not in conflict.split(", ")]
        assignments = ", ".join(f"{column} = excluded.{column}" for column in update_columns)
        return f"{base} on conflict ({conflict}) do update set {assignments};"
    return f"{base};"


def write_sql_files(seed: dict[str, Any]) -> None:
    lines = [
        "-- Destructive seed generated from 选课分享.xlsx and TIC2选课.xlsx.",
        "-- This file deletes existing course data before importing.",
        "-- Execute supabase/schema.sql first, then run this file in Supabase SQL Editor.",
        "begin;",
    ]

    for table in ["reviews", "resources", "course_offerings", "teachers", "courses"]:
        lines.append(f"delete from {table};")

    for course in seed["courses"]:
        lines.append(insert_statement(
            "courses",
            ["id", "name", "category", "audience", "program", "aliases", "summary"],
            [course["id"], course["name"], course["category"], course["audience"], course["program"], course["aliases"], course["summary"]],
        ))

    for teacher in seed["teachers"]:
        lines.append(insert_statement(
            "teachers",
            ["id", "name", "department", "note"],
            [teacher["id"], teacher["name"], teacher["department"], teacher["note"]],
        ))

    for offering in seed["offerings"]:
        lines.append(insert_statement(
            "course_offerings",
            ["id", "course_id", "teacher_id", "term", "status"],
            [offering["id"], offering["courseId"], offering["teacherId"], offering["term"], offering["status"]],
        ))

    for review in seed["reviews"]:
        lines.append(insert_statement(
            "reviews",
            ["id", "offering_id", "term", "workload", "grading", "assessment", "rating", "content", "tags", "contact", "status", "created_at"],
            [review["id"], review["offeringId"], review["term"], review["workload"], review["grading"], review["assessment"], review["rating"], review["content"], review["tags"], review["contact"], review["status"], review["createdAt"]],
        ))

    lines.append("commit;")
    SQL_OUT.write_text("\n".join(lines) + "\n", encoding="utf-8")

    upsert_lines = [
        "-- Non-destructive upsert seed generated from 选课分享.xlsx and TIC2选课.xlsx.",
        "-- This file adds or updates imported courses without deleting existing rows.",
        "-- It only removes obsolete TIC2 course rows created by earlier imports.",
        "begin;",
    ]
    upsert_lines.append(
        "delete from courses where name in ("
        + ", ".join(sql_quote(name) for name in OBSOLETE_TIC2_COURSE_NAMES)
        + ");"
    )
    for course in seed["courses"]:
        upsert_lines.append(insert_statement(
            "courses",
            ["id", "name", "category", "audience", "program", "aliases", "summary"],
            [course["id"], course["name"], course["category"], course["audience"], course["program"], course["aliases"], course["summary"]],
            "id",
        ))
    for teacher in seed["teachers"]:
        upsert_lines.append(insert_statement(
            "teachers",
            ["id", "name", "department", "note"],
            [teacher["id"], teacher["name"], teacher["department"], teacher["note"]],
            "id",
        ))
    for offering in seed["offerings"]:
        upsert_lines.append(insert_statement(
            "course_offerings",
            ["id", "course_id", "teacher_id", "term", "status"],
            [offering["id"], offering["courseId"], offering["teacherId"], offering["term"], offering["status"]],
            "id",
        ))
    for review in seed["reviews"]:
        upsert_lines.append(insert_statement(
            "reviews",
            ["id", "offering_id", "term", "workload", "grading", "assessment", "rating", "content", "tags", "contact", "status", "created_at"],
            [review["id"], review["offeringId"], review["term"], review["workload"], review["grading"], review["assessment"], review["rating"], review["content"], review["tags"], review["contact"], review["status"], review["createdAt"]],
            "id",
        ))
    upsert_lines.append("commit;")
    UPSERT_SQL_OUT.write_text("\n".join(upsert_lines) + "\n", encoding="utf-8")


def main() -> None:
    raw = merge_raw_data(DEFAULT_EXCEL_FILES)
    seed = build_seed(raw)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    body = json.dumps(seed, ensure_ascii=False, indent=2)
    OUT.write_text(
        "import type { SeedData } from '@/lib/types';\n\n"
        f"export const seedData = {body} satisfies SeedData;\n",
        encoding="utf-8",
    )
    write_sql_files(seed)
    print(
        f"Imported {len(seed['courses'])} courses, "
        f"{len(seed['teachers'])} teachers, {len(seed['reviews'])} reviews."
    )


if __name__ == "__main__":
    main()
